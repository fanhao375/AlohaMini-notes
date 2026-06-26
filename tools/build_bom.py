# -*- coding: utf-8 -*-
"""生成 AlohaMini2 BOM 统计 Excel。数据源：AlohaMini/AlohaMini2/docs/BOM.md + AM-ARM/README.md"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"F:\chengshenzhilu\Robot\AlohaMini\AlohaMini2_BOM统计.xlsx"

# ---------- 样式 ----------
HDR_FILL = PatternFill("solid", fgColor="2F5597")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SEC_FILL = PatternFill("solid", fgColor="D6E4F0")
SEC_FONT = Font(bold=True, color="1F3864", size=11)
TOT_FILL = PatternFill("solid", fgColor="FFF2CC")
TOT_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center", vertical="center")
LEF = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIG = Alignment(horizontal="right", vertical="center")

def style_header(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CEN; cell.border = BORDER

def box(ws, r1, r2, ncol):
    for r in range(r1, r2+1):
        for c in range(1, ncol+1):
            ws.cell(row=r, column=c).border = BORDER

wb = openpyxl.Workbook()

# ============ Sheet 1：整机 BOM ============
ws = wb.active; ws.title = "AM2整机BOM"
cols = ["部位","物料","规格","数量","单价¥","小计¥","单价$","采购/备注"]
ws.append(cols); style_header(ws, 1, len(cols))

# (部位, 物料, 规格, 数量, 单价¥, 小计¥, 单价$, 备注)
data = [
("移动底盘","STS-3215 舵机(轮)","ST-3215-C018 12V 1/345",3,110,330,15.97,""),
("移动底盘","STS-3095 舵机(升降)","ST-3095-C002 12V 95kg·cm",1,345,345,50.37,"高扭"),
("移动底盘","全向轮","127mm 74A",3,None,289,40,"¥289/套(3个)"),
("移动底盘","铝型材 T 架","20×40×2 L=400mm",1,10,10,1.5,"或打印件替代"),
("移动底盘","钢销(底盘)","Φ12×80mm",3,3.28,9.84,0.5,"或打印件"),
("移动底盘","钢销(升降轴)","Φ12×25mm",1,3.28,3.28,0.5,"或打印件"),
("移动底盘","轴承(升降导轨)","4×13×5mm",8,3,24,None,""),
("移动底盘","轴承(底盘/臂/升降)","12×18×4mm",8,6,48,None,""),
("移动底盘","摄像头-前","H65V1 720p 2.4mm 1m",1,122,122,17,""),
("移动底盘","摄像头-后","H65V1 720p 2.4mm 1m",1,122,122,17,""),
("移动底盘","摄像头-胸","H65V1 720p 2.4mm 2m",1,122,122,17,""),
("移动底盘","树莓派 5","2GB RAM",1,669,669,93,"Host 算力(可省→USB直连PC)"),
("移动底盘","DC 降压模块","12V→5V 5A PD",1,66,66,9,""),
("移动底盘","散热片","Pi5 被动",1,18,18,2.5,""),
("移动底盘","microSD 卡","32GB",1,99.9,99.9,14,""),
("移动底盘","HDMI→Micro 线","1m 4K60",1,7,7,1,""),
("移动底盘","总线舵机控制器","Waveshare Adapter A",1,27,27,5,""),
("移动底盘","电池 12V","11200mAh DC5521",2,None,114,16,"¥114/对"),
("移动底盘","DC 1分2 线","30cm",1,4.5,4.5,0.6,""),
("移动底盘","DC 延长线","1m",2,2.3,4.6,0.3,""),
("移动底盘","舵机延长线","90cm Feetech 3pin",2,2,4,0.3,""),
("移动底盘","3D 打印件","PLA/PETG/ABS",1,None,None,None,"~4kg 耗材(另计)"),
("Follower双臂","STS-3215 舵机","ST-3215-C018 12V 1/345",8,110,880,15.97,"4/臂"),
("Follower双臂","STS-3095 舵机","ST-3095-C002 12V 95kg·cm",6,345,2070,50.37,"3/臂 高扭"),
("Follower双臂","内六角螺丝 M3×10","100/包",2,None,6,None,"2 包"),
("Follower双臂","热熔螺母","M3×5×4",1,None,5,None,"1 包"),
("Follower双臂","舵机延长线","SCS 3pin 26cm",12,None,3,0.43,"¥3/12根"),
("Follower双臂","摄像头-腕","H65V1 720p 2.4mm 2m",2,122,244,17,"1/臂"),
("Follower双臂","总线舵机控制器","Waveshare Adapter A",2,27,54,5,"1/臂"),
("Follower双臂","USB-C 线","1m 臂→Pi5",4,None,4.8,None,"¥4.8/4根"),
("Follower双臂","3D 打印件","PLA/PETG/ABS",1,None,None,None,"AM-ARM200 ×2"),
("Leader双臂","STS-3215 舵机(leader)","ST-3215-C046 7.4V 1/147",14,110,1540,15.97,"7/臂 可反驱"),
("Leader双臂","热熔螺母","M3×5×4",1,None,5,None,""),
("Leader双臂","总线舵机控制器","Waveshare",2,27,54,12.47,""),
("Leader双臂","电源 5V","DC 5V 适配器",1,None,None,12.99,"无CN价"),
("Leader双臂","USB-C 线","1m 臂→PC",2,10,20,7.19,""),
("Leader双臂","DC 1分2 线","70cm",1,6.5,6.5,None,""),
("Leader双臂","3D 打印件","PLA/PETG/ABS",1,None,None,None,"AM-ARM200 leader ×2"),
("紧固/耗材","热熔螺母 M3","M3×5×4 50/包",1,2.2,2.2,0.3,""),
("紧固/耗材","内六角 M3×10","100/包 蓝胶",3,None,8.5,None,"3 包"),
("紧固/耗材","内六角 M3×18","50/包",1,4.72,4.72,0.65,""),
("紧固/耗材","环氧胶","Loctite E-120HP 50mL",1,74.9,74.9,10,""),
]
r = 2
for row in data:
    ws.append(list(row))
    for c in range(1, len(cols)+1):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = RIG if c in (4,5,6,7) else LEF
    r += 1
last = r-1
ws.freeze_panes = "A2"
widths = [13,22,26,7,9,10,9,26]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width = w

# 合计行（仅累加本表行项目的 小计¥ 列 F2:Flast，避免与汇总混列）
r += 1  # 空一行
ws.cell(row=r, column=1, value="逐项累加合计 (小计¥)").font = TOT_FONT
tcell = ws.cell(row=r, column=6, value="=SUM(F2:F%d)" % last)
tcell.font = TOT_FONT; tcell.fill = TOT_FILL; tcell.alignment = RIG
for c in range(1, 9):
    if c == 6: ws.cell(row=r, column=c).fill = TOT_FILL
    ws.cell(row=r, column=c).border = BORDER
r += 1
ws.cell(row=r, column=1,
        value="官方分项约值：底盘¥2475 / Follower¥3320 / Leader¥1646 / 紧固¥107 = ¥7548 (~$1097)。逐项累加与官方有 ~2% 舍入差；权威分项见『成本汇总』sheet。").font = Font(italic=True, color="666666")

# ============ Sheet 2：舵机统计 ============
ws2 = wb.create_sheet("舵机统计")
h = ["型号","电压","减速比/型号","用途","数量","单价¥","小计¥"]
ws2.append(h); style_header(ws2,1,len(h))
servo = [
("STS-3215","12V","1/345 (C018)","底盘轮 ×3 + Follower 小关节 ×8",11,110,1210),
("STS-3095","12V","95kg·cm (C002)","升降 ×1 + Follower 大关节 ×6 (高扭)",7,345,2415),
("STS-3215","7.4V","1/147 (C046)","Leader 主动臂 ×14 (可反驱)",14,110,1540),
]
r=2
for row in servo:
    ws2.append(list(row))
    for c in range(1,8):
        ws2.cell(row=r,column=c).border=BORDER
        ws2.cell(row=r,column=c).alignment = RIG if c in (5,6,7) else LEF
    r+=1
ws2.cell(row=r,column=1,value="合计").font=TOT_FONT
ws2.cell(row=r,column=5,value=32).font=TOT_FONT
ws2.cell(row=r,column=7,value=5165).font=TOT_FONT
for c in range(1,8):
    ws2.cell(row=r,column=c).fill=TOT_FILL; ws2.cell(row=r,column=c).border=BORDER
    ws2.cell(row=r,column=c).alignment = RIG if c in (5,6,7) else LEF
r+=2
ws2.cell(row=r,column=1,value="舵机占整机 BOM 比例").font=Font(bold=True)
ws2.cell(row=r,column=6,value="¥5,165 / ¥7,548")
ws2.cell(row=r,column=7,value="≈ 68%").font=Font(bold=True,color="C00000")
r+=1
ws2.cell(row=r,column=1,value="机器人本体舵机(去 leader)")
ws2.cell(row=r,column=6,value="¥3,625")
ws2.cell(row=r,column=7,value="18 个")
for i,w in enumerate([12,8,16,42,8,9,10],1): ws2.column_dimensions[get_column_letter(i)].width=w
ws2.freeze_panes="A2"

# ============ Sheet 3：成本汇总 ============
ws3 = wb.create_sheet("成本汇总")
h=["类别","CNY ¥","USD $","占比"]
ws3.append(h); style_header(ws3,1,4)
summ=[("移动底盘",2475,344,"33%"),("Follower 双臂",3320,461,"44%"),
("Leader 双臂(可选)",1646,277,"22%"),("紧固/耗材",107,15,"1%")]
r=2
for row in summ:
    ws3.append(list(row))
    for c in range(1,5):
        ws3.cell(row=r,column=c).border=BORDER
        ws3.cell(row=r,column=c).alignment = RIG if c in (2,3,4) else LEF
    r+=1
ws3.cell(row=r,column=1,value="合计(自打印, 未含耗材)").font=TOT_FONT
ws3.cell(row=r,column=2,value=7548).font=TOT_FONT
ws3.cell(row=r,column=3,value=1097).font=TOT_FONT
ws3.cell(row=r,column=4,value="100%").font=TOT_FONT
for c in range(1,5):
    ws3.cell(row=r,column=c).fill=TOT_FILL; ws3.cell(row=r,column=c).border=BORDER
    ws3.cell(row=r,column=c).alignment = RIG if c in (2,3,4) else LEF
r+=2
notes=[
"关键洞察：",
"• 舵机 ×32 ≈ ¥5,165，占整机 BOM 约 68%（成本由舵机数量决定）。",
"• 机器人本体(去掉 leader 双臂) = ¥5,902 ≈ $820；leader 是遥操作输入设备，可用 VR 手柄替代。",
"• \"<$1000\" 是乐观下限：CN ¥7,548≈$1,036 / US ~$1,097，且未含 ~5kg 耗材(~¥150-300)与常用工具。",
"• Pro 版把部分舵机换 STS-3250 工业舵机，会明显涨价。",
"• 全向轮¥289、电池¥114 按整套/整对计；官方分项为约值，逐项累加有 ~2% 舍入差。",
]
for i,t in enumerate(notes):
    cell=ws3.cell(row=r+i,column=1,value=t)
    if i==0: cell.font=Font(bold=True,size=11)
ws3.column_dimensions["A"].width=22
for col in "BCD": ws3.column_dimensions[col].width=12

# ============ Sheet 4：AM1 vs AM2 ============
ws4 = wb.create_sheet("AM1_vs_AM2")
h=["维度","AlohaMini1","AlohaMini2"]
ws4.append(h); style_header(ws4,1,3)
cmp=[
("臂设计","复用 SO-ARM100 (别人的)","新 AM-ARM200 (自研开源, SO-ARM 衍生)"),
("臂自由度","5+1 DoF","6+1 DoF (+wrist_yaw)"),
("Follower 舵机","12× STS-3215 (全标准)","8× STS-3215 + 6× STS-3095 高扭"),
("Leader 来源","SO-ARM100 leader (别人的)","AM-ARM200 leader (自研)"),
("Leader 舵机","12× STS-3215 7.4V 1/147","14× STS-3215 7.4V 1/147"),
("升降","含底盘 4 舵机内","专用 STS-3095 95kg·cm"),
("臂负载","0.3 kg","1 kg"),
("升降负载","5 kg","30 kg"),
("底盘载荷","10 kg","70 kg"),
("全向轮","4″ ≈100mm","127mm 74A"),
("打印机要求","需大幅面(底盘可能外协)","Bambu P2S 消费级全可打"),
("树莓派","可选(Standalone 才要)","主 BOM 内 (Host/Client WiFi)"),
("USB 直连 PC","官方明确支持(不装单板机)","同样支持"),
]
r=2
for row in cmp:
    ws4.append(list(row))
    for c in range(1,4):
        cell=ws4.cell(row=r,column=c); cell.border=BORDER; cell.alignment=LEF
        if c==1: cell.font=Font(bold=True)
    r+=1
ws4.freeze_panes="A2"
for col,w in zip("ABC",[14,30,38]): ws4.column_dimensions[col].width=w

wb.save(OUT)
print("saved:", OUT)
print("sheets:", wb.sheetnames)
